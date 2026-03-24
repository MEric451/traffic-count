import openpyxl
from openpyxl import load_workbook

def debug_excel_structure(filename):
    """Debug the Excel file structure to understand why direction detection fails"""
    
    wb = load_workbook(filename)
    print(f"Sheet names: {wb.sheetnames}")
    
    # Check the reference sheet (7-8AM)
    if '7-8AM' in wb.sheetnames:
        sheet = wb['7-8AM']
        print(f"\nExamining '7-8AM' sheet:")
        
        # Check first 20 rows in column A
        print("Column A values (first 20 rows):")
        for row in range(1, 21):
            cell_a = sheet.cell(row=row, column=1).value
            cell_b = sheet.cell(row=row, column=2).value
            print(f"  Row {row}: A='{cell_a}' | B='{cell_b}'")
            
            # Check if this could be a direction row
            if cell_a and isinstance(cell_a, str):
                if 'bound' in cell_a.lower() or 'mombasa' in cell_a.lower():
                    print(f"    *** POTENTIAL DIRECTION ROW ***")
    
    # Also check if there are other common sheet names
    common_sheets = ['6-7AM', '8-9AM', '9-10AM', 'DAY']
    for sheet_name in common_sheets:
        if sheet_name in wb.sheetnames:
            print(f"\nSheet '{sheet_name}' exists")
        else:
            print(f"\nSheet '{sheet_name}' NOT FOUND")

if __name__ == "__main__":
    debug_excel_structure("Lukenya Area, Day 1 Monday Counts.xlsx")