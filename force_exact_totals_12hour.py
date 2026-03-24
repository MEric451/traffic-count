import openpyxl
from openpyxl import load_workbook
import random

def force_exact_totals(input_file, output_file, totals):
    """
    Force exact totals to match the image by aggressive distribution (12-hour version)
    """


    wb = load_workbook(input_file)
    
    print("FORCING EXACT TOTALS TO MATCH IMAGE (12-HOUR)")
    print("="*40)
    print(f"All sheets in workbook: {wb.sheetnames}")
    
    # Only process 12-hour sheets (7AM-7PM)
    valid_sheets = ['7-8AM', '8-9AM', '9-10AM', '10-11AM', '11-12PM', '12-1PM', 
                    '1-2PM', '2-3PM', '3-4PM', '4-5PM', '5-6PM', '6-7PM']
    hourly_sheets = [name for name in wb.sheetnames if name in valid_sheets]
    
    print(f"Processing these {len(hourly_sheets)} sheets: {hourly_sheets}")
    
    # Traffic intensity weights for 12-hour distribution (7AM-7PM)
    weights = {
        '7-8AM': 0.10, '8-9AM': 0.12, '9-10AM': 0.09, '10-11AM': 0.08,
        '11-12PM': 0.08, '12-1PM': 0.08, '1-2PM': 0.08, '2-3PM': 0.08,
        '3-4PM': 0.08, '4-5PM': 0.10, '5-6PM': 0.09, '6-7PM': 0.08
    }
    
    # Find direction rows from reference sheet
    reference_sheet = wb['7-8AM']
    direction_rows = {}
    
    for row in range(1, 100):
        cell_a = reference_sheet.cell(row=row, column=1).value
        if cell_a:
            cell_a_clean = ' '.join(str(cell_a).lower().split())  # Normalize spaces
            for direction in totals.keys():
                direction_clean = ' '.join(direction.lower().split())
                if direction_clean in cell_a_clean:
                    cell_b = reference_sheet.cell(row=row, column=2).value
                    if not (isinstance(cell_b, str) and cell_b.startswith('=')):
                        if direction not in direction_rows:
                            direction_rows[direction] = []
                        direction_rows[direction].append(row)
    
    print("\nFound direction rows:")
    for direction, rows in direction_rows.items():
        print(f"  {direction}: {rows}")
    
    # Process each direction and vehicle class
    for direction, direction_totals in totals.items():
        print(f"\nProcessing {direction}...")
        
        if direction not in direction_rows:
            continue
            
        target_rows = direction_rows[direction]
        
        for col_idx, total_12hour in enumerate(direction_totals):
            if total_12hour == 0:
                print(f"  Class {col_idx+1}: Setting to 0 vehicles")
                for sheet_name in hourly_sheets:
                    sheet = wb[sheet_name]
                    for row_num in target_rows:
                        sheet.cell(row=row_num, column=col_idx + 2).value = 0
                continue
                
            print(f"  Class {col_idx+1}: Forcing {total_12hour} vehicles")
            
            # Create hourly distribution with weights
            hourly_distribution = {}
            total_allocated = 0
            
            for sheet_name in hourly_sheets:
                weight = weights.get(sheet_name, 0.0833)  # Default 1/12 if not found
                base_value = total_12hour * weight
                
                # Add variation
                variation = random.uniform(-0.2, 0.2)
                hourly_value = max(1, int(base_value * (1 + variation)))
                hourly_distribution[sheet_name] = hourly_value
                total_allocated += hourly_value
            
            # Force exact total by adjusting
            difference = total_12hour - total_allocated
            
            if difference != 0:
                sheets_list = list(hourly_sheets)
                per_sheet = difference // len(sheets_list)
                remainder = difference % len(sheets_list)
                
                for i, sheet_name in enumerate(sheets_list):
                    adjustment = per_sheet + (1 if i < remainder else 0)
                    hourly_distribution[sheet_name] = max(0, hourly_distribution[sheet_name] + adjustment)
            
            # Verify total is exact
            final_total = sum(hourly_distribution.values())
            if final_total != total_12hour:
                first_sheet = hourly_sheets[0]
                hourly_distribution[first_sheet] += (total_12hour - final_total)
            
            # Apply to sheets
            for sheet_name in hourly_sheets:
                sheet = wb[sheet_name]
                hourly_value = hourly_distribution[sheet_name]
                
                # Distribute across rows in this sheet
                if len(target_rows) > 0 and hourly_value > 0:
                    base_per_row = hourly_value / len(target_rows)
                    remaining = hourly_value
                    
                    for i, row_num in enumerate(target_rows):
                        if i == len(target_rows) - 1:
                            value = remaining
                        else:
                            variation = random.uniform(-0.3, 0.3)
                            value = max(0, int(base_per_row * (1 + variation)))
                            value = min(value, remaining)
                        
                        sheet.cell(row=row_num, column=col_idx + 2).value = value
                        remaining -= value
                else:
                    for row_num in target_rows:
                        sheet.cell(row=row_num, column=col_idx + 2).value = 0
    
    wb.save(output_file)
    print(f"\nSaved to: {output_file}")
    
    verify_forced_totals(wb, totals, hourly_sheets, direction_rows)

def verify_forced_totals(wb, totals, hourly_sheets, direction_rows):
    """Verify the forced totals"""
    print("\nFINAL VERIFICATION:")
    
    for direction, expected_totals in totals.items():
        if direction not in direction_rows:
            continue
            
        target_rows = direction_rows[direction]
        print(f"\n{direction}:")
        
        direction_actual = 0
        direction_expected = sum(expected_totals)
        
        for col_idx, expected_class_total in enumerate(expected_totals):
            class_actual = 0
            
            for sheet_name in hourly_sheets:
                sheet = wb[sheet_name]
                for row_num in target_rows:
                    val = sheet.cell(row=row_num, column=col_idx + 2).value or 0
                    class_actual += val
            
            direction_actual += class_actual
            status = "MATCH" if class_actual == expected_class_total else f"DIFF ({class_actual - expected_class_total})"
            print(f"  Class {col_idx+1}: {expected_class_total} -> {class_actual} [{status}]")
        
        print(f"  TOTAL: {direction_expected} -> {direction_actual}")
        if direction_actual == direction_expected:
            print(f"  [PERFECT MATCH!]")
        else:
            print(f"  [Difference: {direction_actual - direction_expected}]")

if __name__ == "__main__":
    # Define your totals here - change as needed
    totals = {
        "Litein Bound": [1018, 187, 29, 74, 132, 8, 6, 21, 88, 41, 3, 2],
        "Silibwet Bound": [862, 318, 51, 110, 148, 5, 11, 26, 102, 47, 0, 1]
    }

    input_file = "Mogogosiek Site, DAY 1 Monday Litein - Silibwet - Bomet (B124) Road Site Traffic Volumes Hourly Traffic Tally.xlsx"
    output_file = "Mogogosiek Site, DAY 6 Saturday Litein - Silibwet - Bomet (B124) Road Site Traffic Volumes Hourly Traffic Tally.xlsx"
    
    force_exact_totals(input_file, output_file, totals)
