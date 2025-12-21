import openpyxl
from openpyxl import load_workbook
import random

def force_exact_totals(input_file, output_file):
    """
    Force exact totals to match the image by aggressive distribution
    """
    
    # Exact totals from image
    totals = {
        'Bisil Bound': [719, 1799, 954, 543, 551, 34, 9, 58, 677, 269, 363, 9], # Total: 5985
        'Athi River Bound': [726, 1479, 1097, 651, 492, 58, 23, 67, 665, 233, 306, 11] # Total: 5808
    }
    
    wb = load_workbook(input_file)
    hourly_sheets = [name for name in wb.sheetnames if name != 'DAY']
    
    print("FORCING EXACT TOTALS TO MATCH IMAGE BY MUIRURI")
    print("="*40)
    
    # Traffic intensity weights for realistic distribution
    weights = {
        '6-7AM': 0.08, '7-8AM': 0.08, '8-9AM': 0.08, '9-10AM': 0.065, '10-11AM': 0.065,
        '11-12PM': 0.065, '12-1PM': 0.065, '1-2PM': 0.065, '2-3PM': 0.065,
        '3-4PM': 0.065, '4-5PM': 0.08, '5-6PM': 0.08, '6-7PM': 0.08,
        '7-8PM': 0.06, '8-9PM': 0.04, '9-10PM': 0.03, '10-11PM': 0.015,
        '11-12AM': 0.01, '12-1AM': 0.005, '1-2AM': 0.0025, '2-3AM': 0.0025,
        '3-4AM': 0.0025, '4-5AM': 0.0025, '5-6AM': 0.004
    }
    
    # Find direction rows from reference sheet
    reference_sheet = wb['7-8AM']
    direction_rows = {}
    
    for row in range(1, 100):
        cell_a = reference_sheet.cell(row=row, column=1).value
        if cell_a:
            for direction in totals.keys():
                if direction.lower() in str(cell_a).lower():
                    cell_b = reference_sheet.cell(row=row, column=2).value
                    if not (isinstance(cell_b, str) and cell_b.startswith('=')):
                        if direction not in direction_rows:
                            direction_rows[direction] = []
                        direction_rows[direction].append(row)
    
    print("Found direction rows:")
    for direction, rows in direction_rows.items():
        print(f"  {direction}: {rows}")
    
    # Process each direction and vehicle class
    for direction, direction_totals in totals.items():
        print(f"\nProcessing {direction}...")
        
        if direction not in direction_rows:
            continue
            
        target_rows = direction_rows[direction]
        
        for col_idx, total_16hour in enumerate(direction_totals):
            if total_16hour == 0:
                continue
                
            print(f"  Class {col_idx+1}: Forcing {total_16hour} vehicles")
            
            # Create hourly distribution with weights
            hourly_distribution = {}
            total_allocated = 0
            
            for sheet_name in hourly_sheets:
                weight = weights.get(sheet_name, 0.0625)
                base_value = total_16hour * weight
                
                # Add variation
                variation = random.uniform(-0.2, 0.2)
                hourly_value = max(1, int(base_value * (1 + variation)))  # Minimum 1 to ensure distribution
                hourly_distribution[sheet_name] = hourly_value
                total_allocated += hourly_value
            
            # Force exact total by adjusting
            difference = total_16hour - total_allocated
            
            if difference != 0:
                # Distribute difference across all sheets
                sheets_list = list(hourly_sheets)
                per_sheet = difference // len(sheets_list)
                remainder = difference % len(sheets_list)
                
                for i, sheet_name in enumerate(sheets_list):
                    adjustment = per_sheet + (1 if i < remainder else 0)
                    hourly_distribution[sheet_name] = max(0, hourly_distribution[sheet_name] + adjustment)
            
            # Verify total is exact
            final_total = sum(hourly_distribution.values())
            if final_total != total_16hour:
                # Force correction on first sheet
                first_sheet = hourly_sheets[0]
                hourly_distribution[first_sheet] += (total_16hour - final_total)
            
            # Apply to sheets
            for sheet_name in hourly_sheets:
                sheet = wb[sheet_name]
                hourly_value = hourly_distribution[sheet_name]
                
                # Distribute across rows in this sheet
                if len(target_rows) > 0 and hourly_value > 0:
                    base_per_row = hourly_value / len(target_rows)
                    remaining = hourly_value
                    
                    for i, row_num in enumerate(target_rows):
                        if i == len(target_rows) - 1:  # Last row gets remainder
                            value = remaining
                        else:
                            variation = random.uniform(-0.3, 0.3)
                            value = max(0, int(base_per_row * (1 + variation)))
                            value = min(value, remaining)
                        
                        sheet.cell(row=row_num, column=col_idx + 2).value = value
                        remaining -= value
                else:
                    # Set all rows to 0
                    for row_num in target_rows:
                        sheet.cell(row=row_num, column=col_idx + 2).value = 0
    
    wb.save(output_file)
    print(f"\nSaved to: {output_file}")
    
    # Final verification
    verify_forced_totals(wb, totals, hourly_sheets, direction_rows)

def verify_forced_totals(wb, totals, hourly_sheets, direction_rows):
    """Verify the forced totals"""
    print("\nFINAL VERIFICATION:")
    
    for direction, expected_totals in totals.items():
        if direction not in direction_rows:
            continue
            
        target_rows = direction_rows[direction]
        print(f"\n{direction}:")
        
        direction_actual = 0
        direction_expected = sum(expected_totals)
        
        for col_idx, expected_class_total in enumerate(expected_totals):
            if expected_class_total == 0:
                continue
                
            class_actual = 0
            
            for sheet_name in hourly_sheets:
                sheet = wb[sheet_name]
                for row_num in target_rows:
                    val = sheet.cell(row=row_num, column=col_idx + 2).value or 0
                    class_actual += val
            
            direction_actual += class_actual
            status = "MATCH" if class_actual == expected_class_total else f"DIFF ({class_actual - expected_class_total})"
            print(f"  Class {col_idx+1}: {expected_class_total} -> {class_actual} [{status}]")
        
        print(f"  TOTAL: {direction_expected} -> {direction_actual}")
        if direction_actual == direction_expected:
            print(f"  [PERFECT MATCH!]")
        else:
            print(f"  [Difference: {direction_actual - direction_expected}]")

if __name__ == "__main__":
    input_file = "Kitengela Area, Day 3 Wednesday Counts.xlsx"
    output_file = "Kitengela_Forced_Exact_Totals.xlsx"
    
    force_exact_totals(input_file, output_file)