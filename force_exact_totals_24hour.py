import openpyxl
from openpyxl import load_workbook
import random

def force_exact_totals_24hour(input_file, output_file):
    """
    Force exact totals for 24-hour traffic data with realistic distribution
    Also ensures DAY sheet matches image totals exactly
    """
    
    # Exact totals from image (same as 16-hour version)
    totals = {
        'Bisil Bound': [754, 1432, 856, 467, 731, 48, 43, 75, 578, 230, 264, 11], # Total: 5489
        'Athi River Bound': [678, 1201, 934, 522, 642, 56, 34, 71, 561, 214, 282, 14] # Total: 5209
    }
    
    wb = load_workbook(input_file)
    hourly_sheets = [name for name in wb.sheetnames if name != 'DAY']
    
    print("FORCING EXACT TOTALS FOR 24-HOUR DATA BY MUIRURI")
    print("="*50)
    
    # 24-hour traffic intensity weights (your suggested distribution)
    weights = {
        '6-7AM': 0.055, '7-8AM': 0.065, '8-9AM': 0.075, '9-10AM': 0.065, '10-11AM': 0.055,
        '11-12AM': 0.050, '12-1PM': 0.055, '1-2PM': 0.055, '2-3PM': 0.055, '3-4PM': 0.055,
        '4-5PM': 0.065, '5-6PM': 0.075, '6-7PM': 0.070, '7-8PM': 0.055, '8-9PM': 0.045,
        '9-10PM': 0.035, '10-11PM': 0.025, '11-12PM': 0.020, '12-1AM': 0.015, '1-2AM': 0.010,
        '2-3AM': 0.008, '3-4AM': 0.007, '4-5AM': 0.008, '5-6AM': 0.03
    }
    
    # Find direction rows from reference sheet
    reference_sheet = wb['7-8AM']
    direction_rows = {}
    
    for row in range(1, 100):
        cell_a = reference_sheet.cell(row=row, column=1).value
        if cell_a:
            for direction in totals.keys():
                if direction.lower() in str(cell_a).lower():
                    cell_b = reference_sheet.cell(row=row, column=2).value
                    if not (isinstance(cell_b, str) and cell_b.startswith('=')):
                        if direction not in direction_rows:
                            direction_rows[direction] = []
                        direction_rows[direction].append(row)
    
    print("Found direction rows:")
    for direction, rows in direction_rows.items():
        print(f"  {direction}: {rows}")
    
    # Process each direction and vehicle class
    for direction, direction_totals in totals.items():
        print(f"\nProcessing {direction}...")
        
        if direction not in direction_rows:
            continue
            
        target_rows = direction_rows[direction]
        
        for col_idx, total_target in enumerate(direction_totals):
            if total_target == 0:
                continue
                
            print(f"  Class {col_idx+1}: Distributing {total_target} vehicles across 24 hours")
            
            # Create hourly distribution with weights
            hourly_distribution = {}
            total_allocated = 0
            
            for sheet_name in hourly_sheets:
                weight = weights.get(sheet_name, 0.042)  # Default weight for 24 hours
                base_value = total_target * weight
                
                # Add realistic variation
                variation = random.uniform(-0.25, 0.25)
                hourly_value = max(0, int(base_value * (1 + variation)))
                hourly_distribution[sheet_name] = hourly_value
                total_allocated += hourly_value
            
            # Force exact total by adjusting
            difference = total_target - total_allocated
            
            if difference != 0:
                # Distribute difference across peak hours first
                peak_hours = ['7-8AM', '8-9AM', '5-6PM', '6-7PM']
                available_sheets = [s for s in peak_hours if s in hourly_sheets]
                if not available_sheets:
                    available_sheets = hourly_sheets
                
                per_sheet = difference // len(available_sheets)
                remainder = difference % len(available_sheets)
                
                for i, sheet_name in enumerate(available_sheets):
                    adjustment = per_sheet + (1 if i < remainder else 0)
                    hourly_distribution[sheet_name] = max(0, hourly_distribution[sheet_name] + adjustment)
            
            # Final verification and correction
            final_total = sum(hourly_distribution.values())
            if final_total != total_target:
                # Force correction on highest traffic hour
                max_sheet = max(hourly_distribution.keys(), key=lambda x: hourly_distribution[x])
                hourly_distribution[max_sheet] += (total_target - final_total)
            
            # Apply to sheets
            for sheet_name in hourly_sheets:
                sheet = wb[sheet_name]
                hourly_value = hourly_distribution[sheet_name]
                
                # Distribute across rows in this sheet
                if len(target_rows) > 0 and hourly_value > 0:
                    base_per_row = hourly_value / len(target_rows)
                    remaining = hourly_value
                    
                    for i, row_num in enumerate(target_rows):
                        if i == len(target_rows) - 1:  # Last row gets remainder
                            value = remaining
                        else:
                            variation = random.uniform(-0.3, 0.3)
                            value = max(0, int(base_per_row * (1 + variation)))
                            value = min(value, remaining)
                        
                        sheet.cell(row=row_num, column=col_idx + 2).value = value
                        remaining -= value
                else:
                    # Set all rows to 0
                    for row_num in target_rows:
                        sheet.cell(row=row_num, column=col_idx + 2).value = 0
    
    # Fix DAY sheet totals - replace any existing formulas with exact values
    if 'DAY' in wb.sheetnames:
        day_sheet = wb['DAY']
        
        # Clear any existing formulas and set exact totals
        # Find existing direction rows or create new ones
        direction_found = {}
        for row in range(1, 200):
            cell_value = day_sheet.cell(row=row, column=1).value
            if cell_value:
                for direction in totals.keys():
                    if direction.lower() in str(cell_value).lower():
                        direction_found[direction] = row
                        break
        
        # If no existing rows found, create new summary
        if not direction_found:
            start_row = 50
            day_sheet.cell(row=start_row, column=1).value = "DAILY TOTALS SUMMARY"
            direction_found['Bisil Bound'] = start_row + 2
            direction_found['Athi River Bound'] = start_row + 3
            day_sheet.cell(row=start_row+2, column=1).value = "Bisil Bound"
            day_sheet.cell(row=start_row+3, column=1).value = "Athi River Bound"
        
        # Set exact totals (override any formulas)
        for direction, class_totals in totals.items():
            if direction in direction_found:
                row = direction_found[direction]
                for col_idx, total in enumerate(class_totals):
                    day_sheet.cell(row=row, column=col_idx + 2).value = total
                day_sheet.cell(row=row, column=14).value = sum(class_totals)
                print(f"Set DAY sheet {direction}: {sum(class_totals)} total")
    
    wb.save(output_file)
    print(f"\nSaved to: {output_file}")
    
    # Final verification
    verify_24hour_totals(wb, totals, hourly_sheets, direction_rows)

def verify_24hour_totals(wb, totals, hourly_sheets, direction_rows):
    """Verify the 24-hour totals"""
    print("\nFINAL 24-HOUR VERIFICATION:")
    print("="*30)
    
    for direction, expected_totals in totals.items():
        if direction not in direction_rows:
            continue
            
        target_rows = direction_rows[direction]
        print(f"\n{direction}:")
        
        direction_actual = 0
        direction_expected = sum(expected_totals)
        
        for col_idx, expected_class_total in enumerate(expected_totals):
            if expected_class_total == 0:
                continue
                
            class_actual = 0
            
            for sheet_name in hourly_sheets:
                sheet = wb[sheet_name]
                for row_num in target_rows:
                    val = sheet.cell(row=row_num, column=col_idx + 2).value or 0
                    class_actual += val
            
            direction_actual += class_actual
            status = "MATCH" if class_actual == expected_class_total else f"DIFF ({class_actual - expected_class_total})"
            print(f"  Class {col_idx+1}: {expected_class_total} -> {class_actual} [{status}]")
        
        print(f"  TOTAL: {direction_expected} -> {direction_actual}")
        if direction_actual == direction_expected:
            print(f"  [PERFECT 24-HOUR MATCH!]")
        else:
            print(f"  [Difference: {direction_actual - direction_expected}]")

if __name__ == "__main__":
    input_file = "Kitengela Area, Day 3 Wednesday Counts.xlsx"
    output_file = "Kitengela_Perfect_Match.xlsx"
    
    force_exact_totals_24hour(input_file, output_file)