from http.server import BaseHTTPRequestHandler
import json
import base64
import openpyxl
import random
import io
import tempfile
import os

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            
            data = json.loads(post_data.decode('utf-8'))
            file_data = base64.b64decode(data['file'])
            percentage = float(data['percentage'])
            operation = data['operation']
            
            modified_file, log_messages = self.modify_excel(file_data, percentage, operation)
            
            response = {
                'file': base64.b64encode(modified_file).decode('utf-8'),
                'log': '\n'.join(log_messages)
            }
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            error_response = {'error': str(e)}
            self.wfile.write(json.dumps(error_response).encode())
    
    def modify_excel(self, file_data, percentage=13, operation='increase'):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as input_file:
            input_file.write(file_data)
            input_path = input_file.name
        
        wb = openpyxl.load_workbook(input_path)
        
        time_sheets = ['7-8AM', '8-9AM', '9-10AM', '10-11AM', '11-12PM', '12-1PM',
                       '1-2PM', '2-3PM', '3-4PM', '4-5PM', '5-6PM', '6-7PM']
        
        multiplier = (1 + percentage / 100) if operation == 'increase' else (1 - percentage / 100)
        total_modified = 0
        log_messages = []
        
        for sheet_name in time_sheets:
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            sheet_modified = 0
            
            for idx, row in enumerate(ws.iter_rows(), 1):
                cell_a = ws.cell(row=idx, column=1)
                cell_text = str(cell_a.value).lower() if cell_a.value else ''
                
                if not cell_text or cell_text.strip() == '' or cell_text.startswith('='):
                    continue
                    
                has_numeric_data = False
                for col_idx in range(2, 14):
                    cell_value = ws.cell(row=idx, column=col_idx).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        has_numeric_data = True
                        break
                
                if has_numeric_data and len(cell_text.strip()) > 2:
                    for col_idx in range(1, 13):
                        if col_idx < len(row):
                            cell = row[col_idx]
                            if isinstance(cell.value, (int, float)) and not str(cell.value).startswith('='):
                                original = cell.value
                                if original == 0:
                                    continue
                                
                                modified = original * multiplier
                                jitter = 1 + (random.random() * 0.04 - 0.02)
                                new_value = round(modified * jitter)
                                
                                if 1 <= original <= 10:
                                    if operation == 'increase' and new_value <= original:
                                        new_value = original + 1
                                    elif operation == 'decrease' and new_value >= original:
                                        new_value = max(1, original - 1)
                                
                                cell.value = new_value
                                sheet_modified += 1
                                total_modified += 1
            
            log_messages.append(f"Modified {sheet_modified} cells in {sheet_name}")
        
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        os.unlink(input_path)
        
        log_messages.append(f"Total: {total_modified} cells modified")
        
        return output_buffer.getvalue(), log_messages