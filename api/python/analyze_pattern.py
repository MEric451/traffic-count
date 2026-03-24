import openpyxl
from openpyxl import load_workbook
import sys
import json

def analyze_traffic_pattern(excel_file):
    """
    Analyze existing traffic distribution from Excel file
    Returns percentage distribution for each vehicle class and lane ratios
    """
    try:
        wb = load_workbook(excel_file)
        
        # Find DAY sheet or sum from hourly sheets
        if 'DAY' in wb.sheetnames:
            sheet = wb['DAY']
        else:
            # Use first hourly sheet as reference
            hourly_sheets = [s for s in wb.sheetnames if s != 'DAY']
            if not hourly_sheets:
                return {'success': False, 'error': 'No valid sheets found'}
            sheet = wb[hourly_sheets[0]]
        
        # Find lane rows and their totals
        lane_data = {}
        lane_keywords = ['lane', 'bound', 'direction', 'road']
        
        for row in range(1, 100):
            cell_a = sheet.cell(row=row, column=1).value
            
            if cell_a and isinstance(cell_a, str):
                cell_a_lower = cell_a.lower()
                
                # Check if this is a lane row
                if any(keyword in cell_a_lower for keyword in lane_keywords):
                    cell_b = sheet.cell(row=row, column=2).value
                    
                    # Skip formula rows
                    if not (isinstance(cell_b, str) and cell_b.startswith('=')):
                        lane_name = ' '.join(cell_a.split())
                        
                        # Read 12 vehicle class values
                        values = []
                        for col in range(2, 14):  # Columns B to M (12 classes)
                            val = sheet.cell(row=row, column=col).value
                            if isinstance(val, (int, float)):
                                values.append(int(val))
                            else:
                                values.append(0)
                        
                        if sum(values) > 0:  # Only include if has data
                            lane_data[lane_name] = values
        
        if not lane_data:
            return {'success': False, 'error': 'No lane data found'}
        
        # Calculate patterns
        patterns = {}
        grand_total = sum(sum(values) for values in lane_data.values())
        
        for lane_name, values in lane_data.items():
            lane_total = sum(values)
            lane_ratio = lane_total / grand_total if grand_total > 0 else 0
            
            # Calculate percentage for each vehicle class within this lane
            percentages = []
            for val in values:
                pct = (val / lane_total * 100) if lane_total > 0 else 0
                percentages.append(round(pct, 2))
            
            patterns[lane_name] = {
                'total': lane_total,
                'ratio': round(lane_ratio, 3),
                'percentages': percentages,
                'raw_values': values
            }
        
        return {
            'success': True,
            'patterns': patterns,
            'grand_total': grand_total,
            'lane_count': len(lane_data)
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Error analyzing pattern: {str(e)}'
        }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({
            'success': False,
            'error': 'No file path provided'
        }))
        sys.exit(1)
    
    file_path = sys.argv[1]
    result = analyze_traffic_pattern(file_path)
    print(json.dumps(result))
