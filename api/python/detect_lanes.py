import openpyxl
from openpyxl import load_workbook
import sys
import json

def detect_lanes(excel_file):
    """
    Detect lane names from Excel file
    Returns list of lane names found in the Excel
    """
    try:
        wb = load_workbook(excel_file)
        
        # Try to find a reference sheet (prefer 7-8AM, but fallback to first sheet)
        reference_sheet = None
        preferred_sheets = ['7-8AM', '8-9AM', '6-7AM']
        
        for sheet_name in preferred_sheets:
            if sheet_name in wb.sheetnames:
                reference_sheet = wb[sheet_name]
                break
        
        if reference_sheet is None:
            # Use first sheet that's not 'DAY'
            for sheet_name in wb.sheetnames:
                if sheet_name != 'DAY':
                    reference_sheet = wb[sheet_name]
                    break
        
        if reference_sheet is None:
            return {
                'success': False,
                'error': 'No valid sheets found in Excel file'
            }
        
        # Scan column A for lane names
        lanes = []
        lane_keywords = ['lane', 'bound']
        
        for row in range(1, 100):  # Check first 100 rows
            cell_a = reference_sheet.cell(row=row, column=1).value
            
            if cell_a and isinstance(cell_a, str):
                cell_a_lower = cell_a.lower()
                
                # Check if this looks like a lane name
                if any(keyword in cell_a_lower for keyword in lane_keywords):
                    # Check if column B has a number (not a formula)
                    cell_b = reference_sheet.cell(row=row, column=2).value
                    
                    # If column B is not a formula, this is likely a data row
                    if not (isinstance(cell_b, str) and cell_b.startswith('=')):
                        # Clean up the lane name
                        lane_name = ' '.join(cell_a.split())  # Normalize spaces
                        
                        # Avoid duplicates
                        if lane_name not in lanes:
                            lanes.append(lane_name)
        
        if len(lanes) == 0:
            return {
                'success': False,
                'error': 'No lanes detected. Please ensure your Excel has lane names in column A.'
            }
        
        return {
            'success': True,
            'lanes': lanes,
            'sheet_used': reference_sheet.title
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Error reading Excel file: {str(e)}'
        }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({
            'success': False,
            'error': 'No file path provided'
        }))
        sys.exit(1)
    
    file_path = sys.argv[1]
    result = detect_lanes(file_path)
    print(json.dumps(result))
