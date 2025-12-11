import openpyxl
import random
import sys

def modify_excel(input_file, output_file, percentage=13, operation='increase'):
    # Load with data_only to get calculated values
    wb_data = openpyxl.load_workbook(input_file, data_only=True)
    # Load without data_only to preserve formulas
    wb = openpyxl.load_workbook(input_file)
    
    time_sheets = ['7-8AM', '8-9AM', '9-10AM', '10-11AM', '11-12PM', '12-1PM',
                   '1-2PM', '2-3PM', '3-4PM', '4-5PM', '5-6PM', '6-7PM']
    
    # Calculate multiplier based on operation
    multiplier = (1 + percentage / 100) if operation == 'increase' else (1 - percentage / 100)
    
    total_modified = 0
    
    for sheet_name in time_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]
        sheet_modified = 0
        
        for idx, row in enumerate(ws.iter_rows(), 1):
            cell_a_data = ws_data.cell(row=idx, column=1)
            cell_text = str(cell_a_data.value).lower() if cell_a_data.value else ''
            
            # Check if this row contains bound data (has numeric values in columns B-M)
            has_numeric_data = False
            for col_idx in range(1, 13):
                if col_idx < len(row):
                    cell_value = ws_data.cell(row=idx, column=col_idx + 1).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        has_numeric_data = True
                        break
            
            # Only process rows that have both text in column A and numeric data in B-M
            if cell_text and has_numeric_data and len(cell_text.strip()) > 0:
                # Modify columns B-M (indices 1-12)
                for col_idx in range(1, 13):
                    if col_idx < len(row):
                        cell = row[col_idx]
                        if isinstance(cell.value, (int, float)) and not str(cell.value).startswith('='):
                            original = cell.value
                            if original == 0:
                                continue  # Skip zeros
                            
                            modified = original * multiplier
                            jitter = 1 + (random.random() * 0.04 - 0.02)
                            new_value = round(modified * jitter)
                            
                            # Ensure small values change by at least 1
                            if 1 <= original <= 10:
                                if operation == 'increase' and new_value <= original:
                                    new_value = original + 1
                                elif operation == 'decrease' and new_value >= original:
                                    new_value = max(1, original - 1)
                            
                            cell.value = new_value
                            sheet_modified += 1
                            total_modified += 1
        
        print(f"Modified {sheet_modified} cells in {sheet_name}")
    
    wb.save(output_file)
    wb_data.close()
    print(f"\nTotal: {total_modified} cells modified")
    print(f"Saved to: {output_file}")

if __name__ == "__main__":
    input_file = sys.argv[1] if len(sys.argv) > 1 else "Musoli Area Day 2 Tuesday Counts, Sabatia - Bukura (B136) Road Counts.xlsx"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "Modified_Traffic_Counts.xlsx"
    percentage = float(sys.argv[3]) if len(sys.argv) > 3 else 13
    operation = sys.argv[4] if len(sys.argv) > 4 else 'increase'
    modify_excel(input_file, output_file, percentage, operation)
